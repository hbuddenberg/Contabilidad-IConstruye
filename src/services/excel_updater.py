#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel_updater.py

Módulo para copiar el archivo Excel original y actualizarlo con nuevas columnas
de datos extraídos de facturas (Monto Neto, IVA, Total, Estado Subida, URL).

Flujo:
1. Copiar archivo original de "Por Hacer" a "informes"
2. Renombrar con timestamp (fecha_hora)
3. Agregar columnas Q-U con datos de los registros procesados
4. Retornar ruta del archivo actualizado
"""

import os
import shutil
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Definición de columnas nuevas (Q-U)
COLUMNAS_NUEVAS = {
    "Q": "Monto Neto Factura",
    "R": "Monto IVA Factura",
    "S": "Monto Total Factura",
    "T": "Estado Subida Factura",
    "U": "URL Factura",
}

# Índices de columnas (1-based para openpyxl)
COL_MONTO_NETO = 17  # Q
COL_MONTO_IVA = 18  # R
COL_MONTO_TOTAL = 19  # S
COL_ESTADO_SUBIDA = 20  # T
COL_URL_FACTURA = 21  # U


def _generar_nombre_con_timestamp(nombre_original: str, timestamp: str) -> str:
    """
    Genera un nuevo nombre de archivo con timestamp.

    Args:
        nombre_original: Nombre del archivo original (ej: "SEMANA 40 copy.xlsx")
        timestamp: Timestamp de ejecución (ej: "2025-12-08_20.00")

    Returns:
        Nuevo nombre con timestamp (ej: "SEMANA 40 copy_2025-12-08_20.00.xlsx")
    """
    # Separar nombre y extensión
    nombre_base = Path(nombre_original).stem
    extension = Path(nombre_original).suffix

    # Generar nuevo nombre
    nuevo_nombre = f"{nombre_base}_{timestamp}{extension}"

    return nuevo_nombre


def _encontrar_fila_registro(
    df: pd.DataFrame, registro: Any, inicio_datos: int = 2
) -> Optional[int]:
    """
    Encuentra la fila correspondiente a un registro en el DataFrame.

    Busca coincidencia por:
    - Cuenta Proveedor (RUT)
    - Factura (Folio)
    - Fecha documento

    Args:
        df: DataFrame con los datos del Excel
        registro: Objeto Registro con los datos a buscar
        inicio_datos: Fila donde comienzan los datos (1-based, default 2)

    Returns:
        Número de fila (1-based para openpyxl) o None si no encuentra
    """
    try:
        # Crear máscara de búsqueda
        mascara = (
            (
                df["Cuenta Proveedor"].astype(str).str.strip()
                == str(registro.rut_proveedor).strip()
            )
            & (df["Factura"].astype(str).str.strip() == str(registro.folio).strip())
            & (
                df["Fecha documento"].astype(str).str.strip()
                == str(registro.fecha_docto).strip()
            )
        )

        # Obtener índices que coinciden
        indices = df[mascara].index.tolist()

        if indices:
            # Retornar fila (índice + inicio_datos porque índice es 0-based)
            # y la fila 1 es encabezado
            return indices[0] + inicio_datos

        return None

    except Exception as e:
        print(f"⚠️ Error buscando fila para folio {registro.folio}: {e}")
        return None


def _agregar_encabezados(worksheet, fila_encabezado: int = 1):
    """
    Agrega los encabezados de las nuevas columnas (Q-U).

    Args:
        worksheet: Hoja de trabajo de openpyxl
        fila_encabezado: Número de fila para encabezados (default 1)
    """
    # Estilos para encabezados
    font_encabezado = Font(name="Arial", size=11, bold=True)
    fill_encabezado = PatternFill(
        start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
    )
    border_celda = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
    )
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Agregar encabezados
    for col_letra, nombre_columna in COLUMNAS_NUEVAS.items():
        celda = worksheet[f"{col_letra}{fila_encabezado}"]
        celda.value = nombre_columna
        celda.font = font_encabezado
        celda.fill = fill_encabezado
        celda.border = border_celda
        celda.alignment = alignment_center

    # Ajustar ancho de columnas
    worksheet.column_dimensions["Q"].width = 18  # Monto Neto
    worksheet.column_dimensions["R"].width = 18  # Monto IVA
    worksheet.column_dimensions["S"].width = 18  # Monto Total
    worksheet.column_dimensions["T"].width = 20  # Estado Subida
    worksheet.column_dimensions["U"].width = 50  # URL Factura


def _escribir_datos_registro(worksheet, fila: int, registro: Any):
    """
    Escribe los datos de un registro en las columnas Q-U.

    Args:
        worksheet: Hoja de trabajo de openpyxl
        fila: Número de fila donde escribir (1-based)
        registro: Objeto Registro con los datos
    """
    # Estilos
    font_datos = Font(name="Arial", size=10)
    font_url = Font(name="Arial", size=10, color="0563C1", underline="single")
    border_celda = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
    )
    alignment_right = Alignment(horizontal="right")
    alignment_center = Alignment(horizontal="center")

    # Q: Monto Neto Factura
    celda_neto = worksheet.cell(row=fila, column=COL_MONTO_NETO)
    if hasattr(registro, "monto_neto") and registro.monto_neto is not None:
        celda_neto.value = registro.monto_neto
        celda_neto.number_format = "#,##0"
    else:
        celda_neto.value = ""
    celda_neto.font = font_datos
    celda_neto.border = border_celda
    celda_neto.alignment = alignment_right

    # R: Monto IVA Factura
    celda_iva = worksheet.cell(row=fila, column=COL_MONTO_IVA)
    if hasattr(registro, "monto_iva") and registro.monto_iva is not None:
        celda_iva.value = registro.monto_iva
        celda_iva.number_format = "#,##0"
    else:
        celda_iva.value = ""
    celda_iva.font = font_datos
    celda_iva.border = border_celda
    celda_iva.alignment = alignment_right

    # S: Monto Total Factura
    celda_total = worksheet.cell(row=fila, column=COL_MONTO_TOTAL)
    if hasattr(registro, "monto_total") and registro.monto_total is not None:
        celda_total.value = registro.monto_total
        celda_total.number_format = "#,##0"
    else:
        celda_total.value = ""
    celda_total.font = font_datos
    celda_total.border = border_celda
    celda_total.alignment = alignment_right

    # T: Estado Subida Factura
    celda_estado = worksheet.cell(row=fila, column=COL_ESTADO_SUBIDA)
    if hasattr(registro, "estado_subida"):
        if registro.estado_subida is True:
            celda_estado.value = "Subido"
        elif registro.estado_subida is False:
            if hasattr(registro, "error") and registro.error:
                celda_estado.value = f"Error: {registro.error[:30]}"
            else:
                celda_estado.value = "No subido"
        else:
            celda_estado.value = "Pendiente"
    else:
        celda_estado.value = ""
    celda_estado.font = font_datos
    celda_estado.border = border_celda
    celda_estado.alignment = alignment_center

    # U: URL Factura
    celda_url = worksheet.cell(row=fila, column=COL_URL_FACTURA)
    if hasattr(registro, "drive_url") and registro.drive_url:
        celda_url.value = registro.drive_url
        celda_url.font = font_url
        celda_url.hyperlink = registro.drive_url
    else:
        celda_url.value = ""
        celda_url.font = font_datos
    celda_url.border = border_celda


def copiar_y_actualizar_excel(
    ruta_archivo_original: str,
    registros: List[Any],
    directorio_salida: str,
    timestamp: str,
) -> str:
    """
    Copia el archivo Excel original y lo actualiza con las nuevas columnas.

    Flujo:
    1. Copia el archivo original a directorio_salida con nuevo nombre (timestamp)
    2. Abre la copia con openpyxl
    3. Lee los datos con pandas para mapear registros a filas
    4. Agrega encabezados de columnas Q-U
    5. Escribe datos de cada registro en su fila correspondiente
    6. Guarda el archivo

    Args:
        ruta_archivo_original: Ruta completa del archivo original
        registros: Lista de objetos Registro con datos procesados
        directorio_salida: Directorio donde guardar el archivo actualizado
        timestamp: Timestamp para el nombre del archivo (ej: "2025-12-08_20.00")

    Returns:
        Ruta completa del archivo actualizado
    """
    print("\n📊 Iniciando actualización del Excel con datos procesados...")

    # Validar que el archivo original existe
    if not os.path.exists(ruta_archivo_original):
        raise FileNotFoundError(f"Archivo original no existe: {ruta_archivo_original}")

    # Crear directorio de salida si no existe
    os.makedirs(directorio_salida, exist_ok=True)

    # Generar nombre del archivo con timestamp
    nombre_original = os.path.basename(ruta_archivo_original)
    nombre_nuevo = _generar_nombre_con_timestamp(nombre_original, timestamp)
    ruta_archivo_nuevo = os.path.join(directorio_salida, nombre_nuevo)

    print(f"   📄 Archivo original: {nombre_original}")
    print(f"   📄 Archivo nuevo: {nombre_nuevo}")

    # Copiar archivo original al directorio de salida
    print(f"   📋 Copiando archivo a: {directorio_salida}")
    shutil.copy2(ruta_archivo_original, ruta_archivo_nuevo)

    # Leer datos con pandas para mapeo de filas
    print("   📖 Leyendo estructura del archivo...")
    df = pd.read_excel(ruta_archivo_nuevo, dtype=str)

    # Abrir archivo con openpyxl para editar
    print("   ✏️ Abriendo archivo para edición...")
    workbook = openpyxl.load_workbook(ruta_archivo_nuevo)
    worksheet = workbook.active

    # Agregar encabezados de nuevas columnas
    print("   📝 Agregando encabezados de columnas Q-U...")
    _agregar_encabezados(worksheet)

    # Procesar cada registro
    print(f"   🔄 Procesando {len(registros)} registros...")
    registros_actualizados = 0
    registros_no_encontrados = 0

    for registro in registros:
        # Encontrar fila correspondiente
        fila = _encontrar_fila_registro(df, registro)

        if fila is not None:
            # Escribir datos en la fila
            _escribir_datos_registro(worksheet, fila, registro)
            registros_actualizados += 1
        else:
            registros_no_encontrados += 1
            print(
                f"      ⚠️ No se encontró fila para: Folio {registro.folio}, "
                f"RUT {registro.rut_proveedor}"
            )

    # Guardar archivo
    print("   💾 Guardando archivo actualizado...")
    workbook.save(ruta_archivo_nuevo)
    workbook.close()

    # Resumen
    print(f"\n{'=' * 60}")
    print("📊 Actualización de Excel completada:")
    print(f"   ✅ Registros actualizados: {registros_actualizados}")
    print(f"   ⚠️ Registros no encontrados: {registros_no_encontrados}")
    print(f"   📄 Archivo generado: {ruta_archivo_nuevo}")
    print(f"{'=' * 60}\n")

    return ruta_archivo_nuevo


def obtener_resumen_actualizacion(registros: List[Any]) -> Dict[str, Any]:
    """
    Genera un resumen de los datos que se actualizarán.

    Args:
        registros: Lista de registros procesados

    Returns:
        Diccionario con estadísticas de los registros
    """
    total = len(registros)
    con_monto = sum(1 for r in registros if getattr(r, "monto_total", None) is not None)
    subidos = sum(1 for r in registros if getattr(r, "estado_subida", False) is True)
    con_url = sum(1 for r in registros if getattr(r, "drive_url", None) is not None)

    return {
        "total_registros": total,
        "con_montos_extraidos": con_monto,
        "subidos_a_drive": subidos,
        "con_url_drive": con_url,
        "sin_montos": total - con_monto,
        "sin_subir": total - subidos,
    }


# Para pruebas directas
if __name__ == "__main__":
    print("Módulo excel_updater.py")
    print("Este módulo debe ser importado, no ejecutado directamente.")
    print("\nUso:")
    print("  from src.services.excel_updater import copiar_y_actualizar_excel")
    print("  ruta = copiar_y_actualizar_excel(original, registros, salida, timestamp)")
