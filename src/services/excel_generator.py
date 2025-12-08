import os
import sys
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import openpyxl

# Agregar ruta del proyecto para importaciones
sys.path.append(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
)

# Agregar carpeta src al path para importar models
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models.registro import Registro


def generar_informe_excel_con_urls_drive(
    registros_area: List[Registro],
    nombre_area: str,
    directorio_salida: Optional[str] = None,
) -> str:
    """
    Genera un archivo Excel con los registros de un área específica y sus URLs de Drive

    Args:
        registros_area: Lista de objetos Registro
        nombre_area: Nombre del área para el nombre del archivo
        directorio_salida: Directorio donde guardar el archivo (opcional)

    Returns:
        str: Ruta del archivo Excel generado
    """
    # Crear nuevo libro Excel
    libro = openpyxl.Workbook()

    # Crear hoja principal
    hoja_principal = libro.active

    # Crear estilos directamente sin usar NamedStyle para evitar conflictos
    font_encabezado = openpyxl.styles.Font(
        name="Arial", size=12, bold=True, color="FFFFFF"
    )
    border_celda = openpyxl.styles.Border(
        left=openpyxl.styles.Side(border_style="thin", color="000000"),
        right=openpyxl.styles.Side(border_style="thin", color="000000"),
        top=openpyxl.styles.Side(border_style="thin", color="000000"),
        bottom=openpyxl.styles.Side(border_style="thin", color="000000"),
    )
    fill_encabezado = openpyxl.styles.PatternFill(
        start_color="F2F2F2", end_color="DDEBF7", fill_type="solid"
    )
    alignment_center = openpyxl.styles.Alignment(horizontal="center")
    font_datos = openpyxl.styles.Font(name="Arial", size=10)
    font_url = openpyxl.styles.Font(
        name="Arial", size=10, color="0000FF", underline="single"
    )

    # Datos del encabezado
    encabezados = [
        "RUT Proveedor",
        "Razón Social",
        "Folio",
        "Fecha Documento",
        "Área",
        "Estado Drive",
        "Tipo Archivo",
        "Monto Neto",
        "Monto IVA",
        "Monto Total",
        "URL Factura Drive",
    ]

    # Agregar encabezados al Excel
    for col_idx, encabezado in enumerate(encabezados, 1):
        celda = hoja_principal.cell(row=1, column=col_idx)
        celda.value = encabezado
        # Aplicar estilo de encabezado directamente
        celda.font = font_encabezado
        celda.border = border_celda
        celda.fill = fill_encabezado
        celda.alignment = alignment_center

    # Agregar filas de datos
    for fila_idx, registro in enumerate(registros_area, 2):
        # Columna A: RUT Proveedor
        celda = hoja_principal.cell(row=fila_idx, column=1)
        celda.value = registro.rut_proveedor
        celda.font = font_datos
        celda.border = border_celda

        # Columna B: Razón Social
        celda = hoja_principal.cell(row=fila_idx, column=2)
        celda.value = registro.razon_social
        celda.font = font_datos
        celda.border = border_celda

        # Columna C: Folio
        celda = hoja_principal.cell(row=fila_idx, column=3)
        celda.value = registro.folio
        celda.font = font_datos
        celda.border = border_celda

        # Columna D: Fecha Documento
        celda = hoja_principal.cell(row=fila_idx, column=4)
        celda.value = registro.fecha_docto
        celda.font = font_datos
        celda.border = border_celda

        # Columna E: Área
        celda = hoja_principal.cell(row=fila_idx, column=5)
        celda.value = registro.area
        celda.font = font_datos
        celda.border = border_celda

        # Columna F: Estado Drive
        celda = hoja_principal.cell(row=fila_idx, column=6)
        estado = "Subido a Drive" if registro.estado_subida else "No subido"
        celda.value = estado
        celda.font = font_datos
        celda.border = border_celda
        celda.alignment = alignment_center

        # Columna G: Tipo Archivo
        celda = hoja_principal.cell(row=fila_idx, column=7)
        celda.value = registro.tipo_archivo or ""
        celda.font = font_datos
        celda.border = border_celda

        # Columna H: Monto Neto
        celda = hoja_principal.cell(row=fila_idx, column=8)
        if hasattr(registro, "monto_neto") and registro.monto_neto is not None:
            celda.value = registro.monto_neto
            celda.number_format = "#,##0"
        else:
            celda.value = ""
        celda.font = font_datos
        celda.border = border_celda
        celda.alignment = openpyxl.styles.Alignment(horizontal="right")

        # Columna I: Monto IVA
        celda = hoja_principal.cell(row=fila_idx, column=9)
        if hasattr(registro, "monto_iva") and registro.monto_iva is not None:
            celda.value = registro.monto_iva
            celda.number_format = "#,##0"
        else:
            celda.value = ""
        celda.font = font_datos
        celda.border = border_celda
        celda.alignment = openpyxl.styles.Alignment(horizontal="right")

        # Columna J: Monto Total
        celda = hoja_principal.cell(row=fila_idx, column=10)
        if hasattr(registro, "monto_total") and registro.monto_total is not None:
            celda.value = registro.monto_total
            celda.number_format = "#,##0"
        else:
            celda.value = ""
        celda.font = font_datos
        celda.border = border_celda
        celda.alignment = openpyxl.styles.Alignment(horizontal="right")

        # Columna K: URL Factura Drive (CLAVE)
        celda = hoja_principal.cell(row=fila_idx, column=11)
        celda.value = registro.drive_url or ""
        celda.font = font_url
        celda.border = border_celda

    # Ajustar ancho de columnas
    hoja_principal.column_dimensions["A"].width = 18  # RUT
    hoja_principal.column_dimensions["B"].width = 30  # Razón Social
    hoja_principal.column_dimensions["C"].width = 10  # Folio
    hoja_principal.column_dimensions["D"].width = 15  # Fecha Documento
    hoja_principal.column_dimensions["E"].width = 15  # Área
    hoja_principal.column_dimensions["F"].width = 15  # Estado Drive
    hoja_principal.column_dimensions["G"].width = 15  # Tipo Archivo
    hoja_principal.column_dimensions["H"].width = 15  # Monto Neto
    hoja_principal.column_dimensions["I"].width = 15  # Monto IVA
    hoja_principal.column_dimensions["J"].width = 15  # Monto Total
    hoja_principal.column_dimensions["K"].width = 50  # URL Drive

    # Determinar directorio de salida
    if directorio_salida is None:
        directorio_salida = Path.cwd() / "informes"

    # Crear directorio si no existe
    Path(directorio_salida).mkdir(parents=True, exist_ok=True)

    # Generar nombre de archivo
    fecha_hoy = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    nombre_archivo = f"informe_resumen_{nombre_area}_{fecha_hoy}.xlsx"
    ruta_completa = Path(directorio_salida) / nombre_archivo

    # Guardar el archivo
    libro.save(str(ruta_completa))

    print(f"✅ Informe Excel generado: {ruta_completa}")
    return str(ruta_completa)
